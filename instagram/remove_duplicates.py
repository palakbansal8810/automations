import csv
import os
import argparse

INPUT_FILE   = "new_instagram_leads.csv"
OUTPUT_FILE  = "new_instagram_leads.csv"
EXCEL_FILE   = "new_instagram_leads.xlsx"


def deduplicate(input_file=INPUT_FILE, output_file=OUTPUT_FILE, excel_file=EXCEL_FILE):
    if not os.path.isfile(input_file):
        print(f"File not found: {input_file}")
        return

    with open(input_file, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        rows = list(reader)

    if not fieldnames or "username" not in fieldnames:
        print(f"'username' column not found in {input_file}")
        return

    # Add profile_url to fieldnames if not already present
    if "profile_url" not in fieldnames:
        fieldnames = list(fieldnames) + ["profile_url"]

    total = len(rows)
    seen = set()
    unique_rows = []
    duplicates = []

    for row in rows:
        username = row.get("username", "").strip()
        row["profile_url"] = f"http://www.instagram.com/{username}" if username else ""
        if username in seen:
            duplicates.append(username)
        else:
            seen.add(username)
            unique_rows.append(row)

    removed = total - len(unique_rows)

    # Write cleaned CSV
    with open(output_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(unique_rows)

    if removed == 0:
        print(f"No duplicates found. {total} rows are already unique.")
    else:
        print(f"\nDone!")
        print(f"   Total rows before : {total}")
        print(f"   Duplicates removed: {removed}")
        print(f"   Unique rows kept  : {len(unique_rows)}")
        print(f"   CSV saved to      : {output_file}")

        if duplicates:
            print(f"\nRemoved duplicate usernames:")
            for u in duplicates:
                print(f"   - @{u}")

    # Convert to Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"

        # Header row styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="E1306C", end_color="E1306C", fill_type="solid")

        ws.append(fieldnames)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row in unique_rows:
            ws.append([row.get(f, "") for f in fieldnames])

        # Make profile_url cells clickable hyperlinks
        if "profile_url" in fieldnames:
            url_col_idx = fieldnames.index("profile_url") + 1
            url_col_letter = get_column_letter(url_col_idx)
            for row_idx in range(2, len(unique_rows) + 2):
                cell = ws[f"{url_col_letter}{row_idx}"]
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0563C1", underline="single")

        # Auto-fit column widths
        for col_idx, col_name in enumerate(fieldnames, start=1):
            max_len = len(col_name)
            col_letter = get_column_letter(col_idx)
            for row in unique_rows:
                val = row.get(col_name, "")
                if val and len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        wb.save(excel_file)
        print(f"   Excel saved to    : {excel_file}")

    except ImportError:
        print("\nopenpyxl not installed. Installing now...")
        os.system("pip install openpyxl --break-system-packages -q")
        print("Installed. Please run the script again to generate the Excel file.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Deduplicate Instagram leads CSV and export to Excel.")
    parser.add_argument("--input",  default=INPUT_FILE,  help=f"Input CSV file   (default: {INPUT_FILE})")
    parser.add_argument("--output", default=OUTPUT_FILE, help=f"Output CSV file  (default: overwrites input)")
    parser.add_argument("--excel",  default=EXCEL_FILE,  help=f"Output Excel file (default: {EXCEL_FILE})")
    args = parser.parse_args()

    deduplicate(input_file=args.input, output_file=args.output, excel_file=args.excel)